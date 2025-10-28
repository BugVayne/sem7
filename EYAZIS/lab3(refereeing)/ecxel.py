import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo


def create_frame_sheet(workbook, frame_name, slots_data, sheet_type="Фрейм"):
    """
    Создает лист для одного фрейма с таблицей слотов
    """
    # Создаем новый лист
    sheet = workbook.create_sheet(title=frame_name[:31])  # Ограничение длины имени листа

    # Заголовок фрейма
    sheet.merge_cells('A1:D1')
    title_cell = sheet['A1']
    title_cell.value = f"{sheet_type}: {frame_name}"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')

    # Заголовки таблицы
    headers = ['Имя слота', 'Значение слота', 'Способ получения значения', 'Демон']
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=2, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Заполняем данные слотов
    for row, slot in enumerate(slots_data, 3):
        sheet.cell(row=row, column=1).value = slot.get('slot_name', '')
        sheet.cell(row=row, column=2).value = slot.get('value', '')
        sheet.cell(row=row, column=3).value = slot.get('method', '')
        sheet.cell(row=row, column=4).value = slot.get('demon', '')

    # Настраиваем ширину столбцов
    sheet.column_dimensions['A'].width = 25
    sheet.column_dimensions['B'].width = 35
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25

    # Добавляем таблицу с стилем
    max_row = len(slots_data) + 2
    table = Table(displayName=f"Table_{frame_name}", ref=f"A2:D{max_row}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    sheet.add_table(table)


def create_frames_excel():
    """
    Создает Excel файл со всеми фреймами торгового центра
    """
    workbook = openpyxl.Workbook()

    # Удаляем стандартный лист
    workbook.remove(workbook.active)

    # === ФРЕЙМЫ-ПРОТОТИПЫ (ОБЪЕКТЫ И РОЛИ) ===

    # 1. ТорговыйЦентр (прототип)
    torg_center_slots = [
        {'slot_name': 'АКО', 'value': 'Организация', 'method': '-', 'demon': '-'},
        {'slot_name': 'Название', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Адрес', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'КоличествоЭтажей', 'value': '3', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'ОбщаяПлощадь', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'СписокАрендаторов', 'value': '[]', 'method': 'Через присоединенную процедуру',
         'demon': 'IF-ADDED: обновить базу данных'},
        {'slot_name': 'Управляющий', 'value': '-', 'method': 'Через наследование?',
         'demon': 'IF-NEEDED: найти в штатном расписании'},
        {'slot_name': 'Статус', 'value': '"Работает"', 'method': 'По умолчанию',
         'demon': 'IF-CHANGED: уведомить службу безопасности'}
    ]
    create_frame_sheet(workbook, "ТорговыйЦентр", torg_center_slots, "Фрейм-объект (прототип)")

    # 2. СотрудникТЦ (прототип)
    employee_slots = [
        {'slot_name': 'АКО', 'value': 'Человек', 'method': '-', 'demon': '-'},
        {'slot_name': 'Имя', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Должность', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ГрафикРаботы', 'value': '"5/2, 9:00-18:00"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'ТорговыйЦентр', 'value': '-', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "СотрудникТЦ", employee_slots, "Фрейм-роль (прототип)")

    # 3. Арендатор (прототип)
    arendator_slots = [
        {'slot_name': 'АКО', 'value': 'ЮрЛицо', 'method': '-', 'demon': '-'},
        {'slot_name': 'НазваниеМагазина', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТипМагазина', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ПлощадьПомещения', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ДоговорАренды', 'value': '-', 'method': 'Через присоединенную процедуру',
         'demon': 'IF-NEEDED: запросить у юриста'},
        {'slot_name': 'Статус', 'value': '"Активен"', 'method': 'По умолчанию',
         'demon': 'IF-CHANGED: уведомить бухгалтерию'}
    ]
    create_frame_sheet(workbook, "Арендатор", arendator_slots, "Фрейм-роль (прототип)")

    # 4. АрендноеПомещение (прототип)
    room_slots = [
        {'slot_name': 'АКО', 'value': 'Помещение', 'method': '-', 'demon': '-'},
        {'slot_name': 'УникальныйНомер', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Этаж', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Площадь', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТекущийАрендатор', 'value': 'NULL', 'method': 'Через наследование',
         'demon': 'IF-ADDED: внести в реестр ТЦ'}
    ]
    create_frame_sheet(workbook, "АрендноеПомещение", room_slots, "Фрейм-структура (прототип)")

    # === ФРЕЙМЫ-ЭКЗЕМПЛЯРЫ ===

    # 5. ТЦ_Мега (экземпляр)
    tc_mega_slots = [
        {'slot_name': 'АКО', 'value': 'ТорговыйЦентр', 'method': '-', 'demon': '-'},
        {'slot_name': 'Название', 'value': '"МЕГА Белая Дача"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Адрес', 'value': '"Московская обл., г. Котельники"', 'method': 'Из внешних источников',
         'demon': '-'},
        {'slot_name': 'КоличествоЭтажей', 'value': '2', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ОбщаяПлощадь', 'value': '150000', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Управляющий', 'value': 'Сотрудник_Иванов', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "ТЦ_Мега", tc_mega_slots, "Фрейм-экземпляр")

    # 6. Сотрудник_Иванов (экземпляр)
    ivanov_slots = [
        {'slot_name': 'АКО', 'value': 'СотрудникТЦ', 'method': '-', 'demon': '-'},
        {'slot_name': 'Имя', 'value': '"Иванов Алексей Петрович"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Должность', 'value': '"Управляющий ТЦ"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТорговыйЦентр', 'value': 'ТЦ_Мега', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Сотрудник_Иванов", ivanov_slots, "Фрейм-экземпляр")

    # 7. Арендатор_Zara (экземпляр)
    zara_slots = [
        {'slot_name': 'АКО', 'value': 'Арендатор', 'method': '-', 'demon': '-'},
        {'slot_name': 'НазваниеМагазина', 'value': '"Zara"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТипМагазина', 'value': '"Одежда"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ПлощадьПомещения', 'value': '1200', 'method': 'Из внешних источников', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Арендатор_Zara", zara_slots, "Фрейм-экземпляр")

    # 8. Помещение_1_12 (экземпляр)
    room_112_slots = [
        {'slot_name': 'АКО', 'value': 'АрендноеПомещение', 'method': '-', 'demon': '-'},
        {'slot_name': 'УникальныйНомер', 'value': '"1-12"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Этаж', 'value': '1', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Площадь', 'value': '1200', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТекущийАрендатор', 'value': 'Арендатор_Zara', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Помещение_1_12", room_112_slots, "Фрейм-экземпляр")

    # === ФРЕЙМЫ-СИТУАЦИИ ===

    # 9. РабочийДень (прототип ситуации)
    workday_slots = [
        {'slot_name': 'АКО', 'value': 'ШтатнаяСитуация', 'method': '-', 'demon': '-'},
        {'slot_name': 'ТорговыйЦентр', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ВремяНачала', 'value': '"10:00"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'ВремяОкончания', 'value': '"22:00"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'СтатусМагазинов', 'value': '"Открыты"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Посещаемость', 'value': '"Средняя"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Участники', 'value': '[Покупатели, Сотрудники, Арендаторы]', 'method': 'По умолчанию',
         'demon': '-'}
    ]
    create_frame_sheet(workbook, "РабочийДень", workday_slots, "Фрейм-ситуация (прототип)")

    # 10. ЧрезвычайнаяСитуация (прототип)
    emergency_slots = [
        {'slot_name': 'АКО', 'value': 'НештатнаяСитуация', 'method': '-', 'demon': '-'},
        {'slot_name': 'ТорговыйЦентр', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТипЧС', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'УровеньОпасности', 'value': '-', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Действия', 'value': 'Сценарий_Эвакуации', 'method': 'Через наследование', 'demon': '-'},
        {'slot_name': 'Ответственный', 'value': 'СлужбаБезопасности', 'method': 'По умолчанию', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "ЧрезвычайнаяСитуация", emergency_slots, "Фрейм-ситуация (прототип)")

    # 11. ЧС_Пожар_010324 (экземпляр ситуации)
    fire_slots = [
        {'slot_name': 'АКО', 'value': 'ЧрезвычайнаяСитуация', 'method': '-', 'demon': '-'},
        {'slot_name': 'ТорговыйЦентр', 'value': 'ТЦ_Мега', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'ТипЧС', 'value': '"Пожар в фуд-корте"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'УровеньОпасности', 'value': '"Высокий"', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Действия', 'value': 'Сценарий_Эвакуации_ТЦ_Мега', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "ЧС_Пожар_010324", fire_slots, "Фрейм-ситуация (экземпляр)")

    # === ФРЕЙМЫ-СЦЕНАРИИ ===

    # 12. Сценарий_Эвакуации (прототип)
    evacuation_slots = [
        {'slot_name': 'АКО', 'value': 'СценарийБезопасности', 'method': '-', 'demon': '-'},
        {'slot_name': 'Цель', 'value': '"Безопасная эвакуация людей из здания"', 'method': 'По умолчанию',
         'demon': '-'},
        {'slot_name': 'УсловиеЗапуска', 'value': 'ЧрезвычайнаяСитуация.УровеньОпасности = "Высокий"',
         'method': 'По формуле', 'demon': '-'},
        {'slot_name': 'Участники', 'value': '[СлужбаБезопасности, Посетители, Арендаторы]', 'method': 'По умолчанию',
         'demon': '-'},
        {'slot_name': 'Сцены',
         'value': '[Сцена_ОбъявлениеТревоги, Сцена_ВыходКЭвакуационнымВыходам, Сцена_СборНаПлощадке]',
         'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Результат', 'value': '"Все люди эвакуированы"', 'method': 'По умолчанию', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Сценарий_Эвакуации", evacuation_slots, "Фрейм-сценарий (прототип)")

    # 13. Сценарий_Продажи (прототип)
    sale_slots = [
        {'slot_name': 'АКО', 'value': 'БизнесПроцесс', 'method': '-', 'demon': '-'},
        {'slot_name': 'Цель', 'value': '"Оформление покупки товара покупателем"', 'method': 'По умолчанию',
         'demon': '-'},
        {'slot_name': 'УсловиеЗапуска', 'value': '"Покупатель подошел к кассе с товаром"', 'method': 'По умолчанию',
         'demon': '-'},
        {'slot_name': 'Участники', 'value': '[Покупатель, Кассир]', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Сцены', 'value': '[Сцена_СканированиеТоваров, Сцена_Оплата, Сцена_ВыдачаЧека]',
         'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Результат', 'value': '"Покупатель получил товар и чек"', 'method': 'По умолчанию', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Сценарий_Продажи", sale_slots, "Фрейм-сценарий (прототип)")

    # === ФРЕЙМЫ-СЦЕНЫ ===

    # 14. Сцена_ОбъявлениеТревоги (экземпляр)
    alarm_slots = [
        {'slot_name': 'АКО', 'value': 'Сцена', 'method': '-', 'demon': '-'},
        {'slot_name': 'НомерСцены', 'value': '1', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Действие', 'value': '"Включение системы оповещения, передача голосового сообщения"',
         'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Исполнитель', 'value': 'ДежурныйСБ', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Длительность', 'value': '"2 минуты"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'СледующаяСцена', 'value': 'Сцена_ВыходКЭвакуационнымВыходам', 'method': 'Через наследование',
         'demon': '-'}
    ]
    create_frame_sheet(workbook, "Сцена_ОбъявлениеТревоги", alarm_slots, "Фрейм-сцена (экземпляр)")

    # 15. Сцена_Оплата (экземпляр)
    payment_slots = [
        {'slot_name': 'АКО', 'value': 'Сцена', 'method': '-', 'demon': '-'},
        {'slot_name': 'НомерСцены', 'value': '2', 'method': 'Из внешних источников', 'demon': '-'},
        {'slot_name': 'Действие', 'value': '"Оплата товара покупателем (наличные/карта)"', 'method': 'По умолчанию',
         'demon': '-'},
        {'slot_name': 'Исполнитель', 'value': 'Кассир', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'Длительность', 'value': '"1 минута"', 'method': 'По умолчанию', 'demon': '-'},
        {'slot_name': 'СледующаяСцена', 'value': 'Сцена_ВыдачаЧека', 'method': 'Через наследование', 'demon': '-'}
    ]
    create_frame_sheet(workbook, "Сцена_Оплата", payment_slots, "Фрейм-сцена (экземпляр)")

    # Создаем оглавление
    create_table_of_contents(workbook)

    # Сохраняем файл
    workbook.save("ТорговыйЦентр_ФреймоваяМодель.xlsx")
    print("Excel файл 'ТорговыйЦентр_ФреймоваяМодель.xlsx' успешно создан!")


def create_table_of_contents(workbook):
    """
    Создает оглавление с гиперссылками на все фреймы
    """
    # Создаем лист оглавления и перемещаем его на первое место
    toc_sheet = workbook.create_sheet("Оглавление", 0)

    # Заголовок
    toc_sheet.merge_cells('A1:C1')
    title_cell = toc_sheet['A1']
    title_cell.value = "ФРЕЙМОВАЯ МОДЕЛЬ 'ТОРГОВЫЙ ЦЕНТР'"
    title_cell.font = Font(size=16, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_cell.alignment = Alignment(horizontal='center')

    # Заголовки таблицы оглавления
    headers = ['№', 'Название фрейма', 'Тип фрейма']
    for col, header in enumerate(headers, 1):
        cell = toc_sheet.cell(row=3, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    # Список фреймов с типами
    frames_list = [
        (1, "ТорговыйЦентр", "Фрейм-объект (прототип)"),
        (2, "СотрудникТЦ", "Фрейм-роль (прототип)"),
        (3, "Арендатор", "Фрейм-роль (прототип)"),
        (4, "АрендноеПомещение", "Фрейм-структура (прототип)"),
        (5, "ТЦ_Мега", "Фрейм-экземпляр"),
        (6, "Сотрудник_Иванов", "Фрейм-экземпляр"),
        (7, "Арендатор_Zara", "Фрейм-экземпляр"),
        (8, "Помещение_1_12", "Фрейм-экземпляр"),
        (9, "РабочийДень", "Фрейм-ситуация (прототип)"),
        (10, "ЧрезвычайнаяСитуация", "Фрейм-ситуация (прототип)"),
        (11, "ЧС_Пожар_010324", "Фрейм-ситуация (экземпляр)"),
        (12, "Сценарий_Эвакуации", "Фрейм-сценарий (прототип)"),
        (13, "Сценарий_Продажи", "Фрейм-сценарий (прототип)"),
        (14, "Сцена_ОбъявлениеТревоги", "Фрейм-сцена (экземпляр)"),
        (15, "Сцена_Оплата", "Фрейм-сцена (экземпляр)")
    ]

    # Заполняем таблицу оглавления
    for row, (num, frame_name, frame_type) in enumerate(frames_list, 4):
        # Номер
        toc_sheet.cell(row=row, column=1).value = num
        # Название фрейма с гиперссылкой
        frame_cell = toc_sheet.cell(row=row, column=2)
        frame_cell.value = frame_name
        frame_cell.hyperlink = f"#'{frame_name}'!A1"
        frame_cell.font = Font(color="0563C1", underline="single")
        # Тип фрейма
        toc_sheet.cell(row=row, column=3).value = frame_type

    # Настраиваем ширину столбцов
    toc_sheet.column_dimensions['A'].width = 8
    toc_sheet.column_dimensions['B'].width = 30
    toc_sheet.column_dimensions['C'].width = 35

    # Добавляем таблицу с стилем
    table = Table(displayName="Table_Contents", ref=f"A3:C{len(frames_list) + 3}")
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    toc_sheet.add_table(table)


# Запускаем создание Excel файла
if __name__ == "__main__":
    create_frames_excel()